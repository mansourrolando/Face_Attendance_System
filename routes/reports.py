from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from datetime import datetime, date, time as dt_time, timedelta
import pandas as pd
import io
import os
from instance.models import db, Employee, Department, Attendance, Leave, Task, Holiday
from utils.helpers import _calc_work_hours, login_required, get_setting, format_work_hours, format_minutes, log_action
from utils.csrf import validate_csrf_token
from utils.absence_utils import mark_absent_employees, is_non_working_day
from config import *

reports_bp = Blueprint('reports', __name__)


def _calculate_column_widths(worksheet, df, min_width=12, max_width=45):
    """حساب عرض الأعمدة تلقائياً بناءً على المحتوى - يدعم النصوص العربية"""
    from openpyxl.utils import get_column_letter

    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = 0
        # حساب طول اسم العمود (العنوان)
        header_len = 0
        for ch in str(col_name):
            if '\u0600' <= ch <= '\u06FF' or '\uFB50' <= ch <= '\uFEFF':
                header_len += 1.6  # الأحرف العربية أعرض
            elif ch.isupper():
                header_len += 1.2
            else:
                header_len += 1.0
        max_length = header_len

        # حساب أطول قيمة في العمود
        for row_idx in range(len(df)):
            cell_value = df.iloc[row_idx, col_idx - 1]
            cell_str = str(cell_value) if cell_value is not None else ''
            cell_len = 0
            for ch in cell_str:
                if '\u0600' <= ch <= '\u06FF' or '\uFB50' <= ch <= '\uFEFF':
                    cell_len += 1.6  # الأحرف العربية أعرض
                elif ch.isupper():
                    cell_len += 1.2
                else:
                    cell_len += 1.0
            if cell_len > max_length:
                max_length = cell_len

        # إضافة padding وتطبيق الحدود
        calculated_width = max_length + 4
        final_width = max(min_width, min(calculated_width, max_width))
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = final_width

        # ارتفاع صف العنوان
        worksheet.row_dimensions[1].height = 30


def _auto_mark_range(start_d, end_d):
    """تسجيل غياب تلقائي لنطاق تواريخ"""
    today = date.today()
    now = datetime.now()

    work_end_setting = get_setting('work_end', WORK_END_TIME)
    try:
        end_hour, end_minute = map(int, work_end_setting.split(':'))
    except:
        end_hour, end_minute = map(int, WORK_END_TIME.split(':'))
    work_ended_today = now.time() >= dt_time(end_hour, end_minute)

    current_d = start_d
    while current_d <= end_d:
        if current_d < today:
            mark_absent_employees(current_d)
        elif current_d == today and work_ended_today:
            mark_absent_employees(current_d, allow_today=True)
        current_d += timedelta(days=1)


def get_filtered_records(request_args, limit=None):
    """
    دالة مشتركة لجلب السجلات المفلترة - تستخدم في التقارير والتصدير
    تعيد: (unified_records, leave_types, summary, employees, departments)
    """
    start_date_str = request_args.get('start_date')
    end_date_str = request_args.get('end_date')
    filter_employee_id = request_args.get('employee_id')
    filter_department_id = request_args.get('department_id')
    filter_status = request_args.get('status')
    filter_emp_status = request_args.get('emp_status')

    start_date = None
    end_date = None

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except:
            pass

    # تسجيل غياب تلقائي
    today = date.today()
    auto_absent_start = start_date if start_date else today - timedelta(days=30)
    auto_absent_end = end_date if end_date else today
    _auto_mark_range(auto_absent_start, auto_absent_end)

    # تحديد نوع الفلتر
    is_task_filter = filter_status in ('admin_task', 'overtime')
    is_leave_type_filter = filter_status and filter_status.startswith('leave_')
    is_absence_type_filter = filter_status in ('absent_justified', 'absent_unjustified')
    show_attendance = not is_task_filter
    show_tasks = not filter_status or is_task_filter

    # === جلب سجلات الحضور ===
    attendance_records = []
    if show_attendance:
        query = Attendance.query

        if start_date:
            query = query.filter(Attendance.date >= start_date)
        if end_date:
            query = query.filter(Attendance.date <= end_date)
        if filter_employee_id:
            query = query.filter(Attendance.employee_id == filter_employee_id)

        if filter_status and not is_task_filter:
            if is_absence_type_filter:
                query = query.filter(Attendance.status == 'absent', Attendance.absence_type == filter_status.replace('absent_', ''))
            elif is_leave_type_filter:
                query = query.filter(Attendance.status == 'leave')
            else:
                query = query.filter(Attendance.status == filter_status)

        _joined_employee = False
        if filter_department_id:
            query = query.join(Employee).filter(Employee.department_id == filter_department_id)
            _joined_employee = True

        if filter_emp_status:
            if not _joined_employee:
                query = query.join(Employee)
                _joined_employee = True
            query = query.filter(Employee.status == filter_emp_status)

        if not _joined_employee:
            query = query.join(Employee)

        query = query.filter(
            db.or_(
                Employee.hire_date == None,
                Attendance.date >= Employee.hire_date
            )
        )

        if limit:
            attendance_records = query.order_by(Attendance.date.desc()).limit(limit).all()
        else:
            attendance_records = query.order_by(Attendance.date.desc()).all()

        # فلترة حسب نوع الإجازة المحدد
        if is_leave_type_filter:
            leave_type_val = filter_status.replace('leave_', '')
            filtered = []
            for r in attendance_records:
                if r.status == 'leave' and r.employee_id:
                    leave_rec = Leave.query.filter(
                        Leave.employee_id == r.employee_id,
                        Leave.start_date <= r.date,
                        Leave.end_date >= r.date,
                        Leave.status == 'approved'
                    ).first()
                    if leave_rec and leave_rec.leave_type == leave_type_val:
                        filtered.append(r)
                else:
                    filtered.append(r)
            attendance_records = filtered

    # === جلب سجلات المهام والعمل الإضافي ===
    task_records = []
    if show_tasks:
        task_query = Task.query

        if start_date:
            task_query = task_query.filter(Task.date >= start_date)
        if end_date:
            task_query = task_query.filter(Task.date <= end_date)
        if filter_employee_id:
            task_query = task_query.filter(Task.employee_id == filter_employee_id)
        if is_task_filter:
            task_query = task_query.filter(Task.task_type == filter_status)
        if filter_department_id:
            task_query = task_query.join(Employee).filter(Employee.department_id == filter_department_id)

        if filter_emp_status:
            if not filter_department_id:
                task_query = task_query.join(Employee)
            task_query = task_query.filter(Employee.status == filter_emp_status)

        task_records = task_query.order_by(Task.date.desc()).all()

    # === توحيد السجلات ===
    unified_records = []

    for r in attendance_records:
        hourly_leave_duration = None
        if r.leave_out and r.leave_in:
            from datetime import datetime as dt_mod
            leave_diff = dt_mod.combine(r.date, r.leave_in) - dt_mod.combine(r.date, r.leave_out)
            leave_minutes = int(leave_diff.total_seconds() / 60)
            lh = leave_minutes // 60
            lm = leave_minutes % 60
            hourly_leave_duration = f'{lh}س {lm}د'

        unified_records.append({
            'id': r.id,
            'record_type': 'attendance',
            'employee': r.employee,
            'date': r.date,
            'time_in': r.time_in,
            'time_out': r.time_out,
            'leave_out': r.leave_out,
            'leave_in': r.leave_in,
            'hourly_leave_duration': hourly_leave_duration,
            'status': r.status,
            'absence_type': r.absence_type,
            'total_hours': r.work_hours,
            'work_hours_display': format_work_hours(r.work_hours) if r.work_hours else _calc_work_hours(r.time_in, r.time_out),
            'late_minutes': r.late_minutes or 0,
            'early_checkout': r.early_checkout or False,
            'early_checkout_minutes': r.early_checkout_minutes or 0,
            'description': r.notes or '',
        })

    for t in task_records:
        unified_records.append({
            'id': t.id,
            'record_type': 'task',
            'employee': t.employee,
            'date': t.date,
            'time_in': t.start_time,
            'time_out': t.end_time,
            'status': t.task_type,
            'absence_type': None,
            'total_hours': t.total_hours,
            'work_hours_display': str(t.total_hours) + 'س' if t.total_hours else '-',
            'late_minutes': 0,
            'early_checkout': False,
            'early_checkout_minutes': 0,
            'description': t.description or '',
        })

    unified_records.sort(key=lambda x: x['date'], reverse=True)

    # === جلب أنواع الإجازات ===
    leave_types = {}
    for r in unified_records:
        if r['record_type'] == 'attendance' and r['status'] == 'leave' and r['employee']:
            leave_rec = Leave.query.filter(
                Leave.employee_id == r['employee'].id,
                Leave.start_date <= r['date'],
                Leave.end_date >= r['date'],
                Leave.status == 'approved'
            ).first()
            if leave_rec:
                leave_types[f"att_{r['id']}"] = leave_rec.leave_type

    # === حساب الملخص ===
    summary = {
        'early': 0,
        'present': sum(1 for r in unified_records if r['record_type'] == 'attendance' and r['status'] == 'present'),
        'late': sum(1 for r in unified_records if r['record_type'] == 'attendance' and r['status'] == 'late'),
        'leave': sum(1 for r in unified_records if r['record_type'] == 'attendance' and r['status'] == 'leave'),
        'absent_justified': sum(1 for r in unified_records if r['record_type'] == 'attendance' and r['status'] == 'absent' and r.get('absence_type') == 'justified'),
        'absent_unjustified': sum(1 for r in unified_records if r['record_type'] == 'attendance' and r['status'] == 'absent' and r.get('absence_type') != 'justified'),
        'admin_task': sum(1 for r in unified_records if r['record_type'] == 'task' and r['status'] == 'admin_task'),
        'overtime': sum(1 for r in unified_records if r['record_type'] == 'task' and r['status'] == 'overtime'),
        'total_work_hours': sum(r.get('total_hours', 0) or 0 for r in unified_records if r['record_type'] == 'attendance'),
        'total_late_minutes': sum(r.get('late_minutes', 0) or 0 for r in unified_records if r['record_type'] == 'attendance'),
        'total_early_checkout': sum(1 for r in unified_records if r.get('early_checkout')),
    }

    employees = Employee.query.all()
    departments = Department.query.all()

    return unified_records, leave_types, summary, employees, departments


@reports_bp.route('/reports', endpoint='reports')
@login_required()
def reports():
    unified_records, leave_types, summary, employees, departments = get_filtered_records(
        request.args, limit=500
    )
    return render_template('reports.html',
        records=unified_records, summary=summary,
        employees=employees, departments=departments,
        leave_types=leave_types
    )


@reports_bp.route('/reports/print', endpoint='reports_print')
@login_required()
def reports_print():
    """عرض التقارير للطباعة - بدون شريط جانبي"""
    unified_records, leave_types, summary, employees, departments = get_filtered_records(
        request.args
    )
    return render_template('reports_print.html',
        records=unified_records, summary=summary,
        employees=employees, departments=departments,
        leave_types=leave_types,
        print_date=datetime.now().strftime('%Y-%m-%d %H:%M')
    )


@reports_bp.route('/export_reports', endpoint='export_reports')
@login_required()
def export_reports():
    """تصدير التقارير إلى Excel"""
    unified_records, leave_types, summary, employees, departments = get_filtered_records(
        request.args
    )

    if not unified_records:
        flash('لا توجد سجلات مطابقة للتصدير', 'warning')
        return redirect(url_for('reports.reports', **request.args))

    leave_type_map = {'paid': 'مأجورة', 'unpaid': 'بلا راتب', 'medical': 'صحية', 'maternity': 'أمومة', 'hourly': 'ساعية'}

    status_map = {
        'present': 'حاضر', 'late': 'متأخر',
        'leave': 'إجازة', 'absent': 'غائب',
        'admin_task': 'مهمة إدارية', 'overtime': 'عمل إضافي',
    }
    status_colors = {
        'present': '27AE60', 'late': 'E67E22', 'leave': '3498DB', 'absent': 'E74C3C',
        'absent_justified': 'E67E22', 'absent_unjustified': 'C0392B',
        'admin_task': '8E44AD', 'overtime': 'F39C12',
    }

    data = []
    status_raw_list = []

    for record in unified_records:
        status_text = status_map.get(record['status'], record['status'])
        leave_type_text = ''
        task_type_text = ''
        absence_type_text = ''

        if record['record_type'] == 'attendance' and record['status'] == 'leave':
            lt_key = f"att_{record['id']}"
            if lt_key in leave_types:
                lt = leave_types[lt_key]
                leave_type_text = leave_type_map.get(lt, lt)
                status_text = f'إجازة ({leave_type_text})'

        if record['record_type'] == 'attendance' and record['status'] == 'absent':
            abs_type = record.get('absence_type')
            if abs_type == 'justified':
                absence_type_text = 'مبرر'
                status_text = 'غياب مبرر'
            else:
                absence_type_text = 'غير مبرر'
                status_text = 'غياب غير مبرر'

        if record['record_type'] == 'task':
            task_type_text = 'مهمة إدارية' if record['status'] == 'admin_task' else 'عمل إضافي'

        color_key = record['status']
        if record['status'] == 'absent':
            color_key = f"absent_{record.get('absence_type', 'unjustified')}"

        data.append({
            'الموظف': record['employee'].name if record['employee'] else '-',
            'رقم الموظف': record['employee'].employee_id if record['employee'] else '-',
            'القسم': record['employee'].department.name if record['employee'] and record['employee'].department else '-',
            'حالة الموظف': 'نشط' if record['employee'] and record['employee'].status == 'active' else 'غير نشط',
            'التاريخ': str(record['date']),
            'من': record['time_in'].strftime('%H:%M:%S') if record['time_in'] else '-',
            'إلى': record['time_out'].strftime('%H:%M:%S') if record['time_out'] else '-',
            'مدة العمل': record['work_hours_display'] if record.get('work_hours_display') else _calc_work_hours(record['time_in'], record['time_out']),
            'دقائق التأخير': record.get('late_minutes', 0),
            'خروج مبكر': 'نعم' if record.get('early_checkout') else 'لا',
            'دقائق الخروج المبكر': record.get('early_checkout_minutes', 0),
            'خروج إجازة': record['leave_out'].strftime('%H:%M') if record.get('leave_out') else '-',
            'رجوع إجازة': record['leave_in'].strftime('%H:%M') if record.get('leave_in') else '-',
            'مدة الإجازة': record.get('hourly_leave_duration', '-'),
            'الحالة': status_text,
            'نوع الإجازة': leave_type_text if leave_type_text else '-',
            'نوع الغياب': absence_type_text if absence_type_text else '-',
            'النوع': task_type_text if task_type_text else '-',
            'الوصف': record['description'] if record['description'] else '-',
        })
        status_raw_list.append(color_key)

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='تقرير الحضور')

        workbook = writer.book
        worksheet = writer.sheets['تقرير الحضور']

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC'),
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_font = Font(name='Arial', size=11)
        row_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        status_col = df.columns.get_loc('الحالة') + 1

        for row_num in range(2, len(df) + 2):
            row_fill = even_fill if row_num % 2 == 0 else odd_fill
            status_raw = status_raw_list[row_num - 2]
            worksheet.row_dimensions[row_num].height = 22

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = row_font
                cell.alignment = row_alignment
                cell.border = thin_border
                cell.fill = row_fill

            status_cell = worksheet.cell(row=row_num, column=status_col)
            color = status_colors.get(status_raw, '95A5A6')
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            status_cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

        # حساب عرض الأعمدة تلقائياً
        _calculate_column_widths(worksheet, df)

        worksheet.freeze_panes = 'A2'

        # ===== صف الملخص =====
        summary_row = len(df) + 3
        summary_font = Font(name='Arial', bold=True, size=12, color='2C3E50')
        summary_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')

        worksheet.cell(row=summary_row, column=1, value='ملخص التقرير').font = Font(name='Arial', bold=True, size=13, color='2C3E50')
        worksheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)

        summary_data = [
            ('إجمالي الحضور', summary['present'], '27AE60'),
            ('إجمالي التأخير', summary['late'], 'E67E22'),
            ('إجمالي الإجازات', summary['leave'], '3498DB'),
            ('إجمالي الغياب المبرر', summary['absent_justified'], 'E67E22'),
            ('إجمالي الغياب غير المبرر', summary['absent_unjustified'], 'C0392B'),
            ('إجمالي المهام الإدارية', summary['admin_task'], '8E44AD'),
            ('إجمالي العمل الإضافي', summary['overtime'], 'F39C12'),
            ('إجمالي ساعات العمل', f"{summary['total_work_hours']:.1f} ساعة", '2C3E50'),
            ('إجمالي دقائق التأخير', format_minutes(summary['total_late_minutes']), 'E67E22'),
            ('إجمالي حالات الخروج المبكر', summary['total_early_checkout'], 'C0392B'),
            ('إجمالي السجلات', len(unified_records), '2C3E50'),
        ]

        for i, (label, value, color) in enumerate(summary_data):
            row = summary_row + 1 + i
            label_cell = worksheet.cell(row=row, column=1, value=label)
            label_cell.font = summary_font
            label_cell.fill = summary_fill
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            label_cell.border = thin_border
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

            value_cell = worksheet.cell(row=row, column=3, value=str(value))
            value_cell.font = Font(name='Arial', bold=True, size=12, color=color)
            value_cell.fill = summary_fill
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.border = thin_border

        worksheet.sheet_view.rightToLeft = True

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'attendance_report_{date.today()}.xlsx'
    )


@reports_bp.route('/export_payroll', endpoint='export_payroll')
@login_required()
def export_payroll():
    """تصدير بيانات الرواتب - ملف Excel بتنسيق مناسب للرواتب"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    filter_department_id = request.args.get('department_id')

    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except:
            pass

    if not start_date:
        today = date.today()
        start_date = date(today.year, today.month, 1)
    if not end_date:
        end_date = date.today()

    # جلب الموظفين النشطين
    emp_query = Employee.query.filter_by(status='active')
    if filter_department_id:
        emp_query = emp_query.filter_by(department_id=filter_department_id)
    employees = emp_query.order_by(Employee.name).all()

    payroll_data = []
    for emp in employees:
        # جلب سجلات الحضور للفترة
        records = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).all()

        present_days = sum(1 for r in records if r.status == 'present')
        late_days = sum(1 for r in records if r.status == 'late')
        absent_days = sum(1 for r in records if r.status == 'absent')
        leave_days = sum(1 for r in records if r.status == 'leave')
        total_work_hours = sum(r.work_hours or 0 for r in records)
        total_late_minutes = sum(r.late_minutes or 0 for r in records)
        early_checkouts = sum(1 for r in records if r.early_checkout)

        # العمل الإضافي
        overtime_records = Task.query.filter(
            Task.employee_id == emp.id,
            Task.task_type == 'overtime',
            Task.date >= start_date,
            Task.date <= end_date
        ).all()
        overtime_hours = sum(t.total_hours or 0 for t in overtime_records)

        # رصيد الإجازات
        annual_balance = emp.annual_leave_balance or 0
        sick_balance = emp.sick_leave_balance or 0

        payroll_data.append({
            'رقم الموظف': emp.employee_id,
            'اسم الموظف': emp.name,
            'القسم': emp.department.name if emp.department else '-',
            'المسمى الوظيفي': emp.position or '-',
            'أيام الحضور': present_days,
            'أيام التأخير': late_days,
            'أيام الغياب': absent_days,
            'أيام الإجازة': leave_days,
            'إجمالي ساعات العمل': round(total_work_hours, 1),
            'إجمالي دقائق التأخير': total_late_minutes,
            'عدد مرات الخروج المبكر': early_checkouts,
            'ساعات العمل الإضافي': overtime_hours,
            'رصيد الإجازات السنوية': annual_balance,
            'رصيد الإجازات المرضية': sick_balance,
        })

    df = pd.DataFrame(payroll_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='بيانات الرواتب')

        workbook = writer.book
        worksheet = writer.sheets['بيانات الرواتب']

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC'),
        )

        # تنسيق صف العنوان
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        worksheet.row_dimensions[1].height = 30

        # تنسيق صفوف البيانات
        row_font = Font(name='Arial', size=11)
        row_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for row_num in range(2, len(df) + 2):
            row_fill = even_fill if row_num % 2 == 0 else odd_fill
            worksheet.row_dimensions[row_num].height = 22
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = row_font
                cell.alignment = row_alignment
                cell.border = thin_border
                cell.fill = row_fill

        # حساب عرض الأعمدة تلقائياً
        _calculate_column_widths(worksheet, df)

        worksheet.sheet_view.rightToLeft = True

        # إضافة معلومات الفترة
        info_row = len(df) + 3
        worksheet.cell(row=info_row, column=1, value=f'الفترة: من {start_date} إلى {end_date}').font = Font(name='Arial', size=10, color='7F8C8D', italic=True)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'payroll_data_{start_date}_{end_date}.xlsx'
    )


@reports_bp.route('/export_employees', endpoint='export_employees')
@login_required()
def export_employees():
    """تصدير بيانات الموظفين إلى Excel"""
    employees = Employee.query.order_by(Employee.id).all()

    data = []
    for emp in employees:
        current_month = date.today().month
        current_year = date.today().year

        month_present = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.status == 'present',
            db.extract('month', Attendance.date) == current_month,
            db.extract('year', Attendance.date) == current_year
        ).count()

        month_late = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.status == 'late',
            db.extract('month', Attendance.date) == current_month,
            db.extract('year', Attendance.date) == current_year
        ).count()

        month_leave = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.status == 'leave',
            db.extract('month', Attendance.date) == current_month,
            db.extract('year', Attendance.date) == current_year
        ).count()

        month_absent = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.status == 'absent',
            db.extract('month', Attendance.date) == current_month,
            db.extract('year', Attendance.date) == current_year
        ).count()

        data.append({
            'الرقم الوظيفي': emp.employee_id,
            'الاسم': emp.name,
            'البريد': emp.email or '-',
            'الهاتف': emp.phone or '-',
            'الوظيفة': emp.position or '-',
            'القسم': emp.department.name if emp.department else '-',
            'تاريخ التوظيف': emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '-',
            'الحالة': 'نشط' if emp.status == 'active' else 'غير نشط',
            'بصمة الوجه': 'مسجلة' if emp.face_encoding else 'غير مسجلة',
            'رصيد الإجازات السنوية': emp.annual_leave_balance or 0,
            'رصيد الإجازات المرضية': emp.sick_leave_balance or 0,
            'حضور الشهر': month_present,
            'تأخير الشهر': month_late,
            'إجازات الشهر': month_leave,
            'غياب الشهر': month_absent,
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='بيانات الموظفين')

        workbook = writer.book
        worksheet = writer.sheets['بيانات الموظفين']

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC'),
        )

        # تنسيق صف العنوان
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        worksheet.row_dimensions[1].height = 30

        # تنسيق صفوف البيانات
        row_font = Font(name='Arial', size=11)
        row_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # تحديد عمود الحالة للألوان
        status_col_idx = None
        for ci, col_name in enumerate(df.columns, 1):
            if col_name == 'الحالة':
                status_col_idx = ci
                break

        for row_num in range(2, len(df) + 2):
            row_fill = even_fill if row_num % 2 == 0 else odd_fill
            worksheet.row_dimensions[row_num].height = 22
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = row_font
                cell.alignment = row_alignment
                cell.border = thin_border
                cell.fill = row_fill

            # تلوين خلية الحالة
            if status_col_idx:
                status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                status_val = str(status_cell.value) if status_cell.value else ''
                if 'نشط' in status_val:
                    status_cell.fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
                    status_cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                elif 'غير' in status_val:
                    status_cell.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
                    status_cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

        # حساب عرض الأعمدة تلقائياً
        _calculate_column_widths(worksheet, df)

        worksheet.sheet_view.rightToLeft = True

    output.seek(0)

    log_action('read', 'employee', description='تصدير بيانات الموظفين إلى Excel')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'employees_data_{date.today()}.xlsx'
    )


@reports_bp.route('/import_employees', methods=['GET', 'POST'], endpoint='import_employees')
@login_required()
def import_employees():
    """استيراد موظفين من ملف Excel"""
    if request.method == 'POST':
        if not validate_csrf_token(request.form.get('csrf_token')):
            flash('رمز الأمان غير صالح، حاول مرة أخرى', 'danger')
            return redirect(request.url)

        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('يرجى اختيار ملف Excel (.xlsx أو .xls)', 'danger')
            return redirect(request.url)

        try:
            df = pd.read_excel(file)

            # البحث عن عمود الاسم - يدعم أسماء أعمدة متعددة
            name_col = None
            name_col_aliases = ['الاسم', 'اسم الموظف', 'الموظف', 'name', 'employee', 'employee name', 'اسم']
            for col in df.columns:
                col_clean = str(col).strip()
                if col_clean in name_col_aliases or col_clean.lower() in [a.lower() for a in name_col_aliases]:
                    name_col = col
                    break

            if name_col is None:
                flash('يجب أن يحتوي الملف على عمود "الاسم" أو "الموظف" أو "name"', 'danger')
                return redirect(request.url)

            # البحث عن أعمدة اختيارية
            email_col = None
            for col in df.columns:
                col_clean = str(col).strip().lower()
                if col_clean in ['البريد', 'البريد الإلكتروني', 'email', 'e-mail']:
                    email_col = col
                    break

            phone_col = None
            for col in df.columns:
                col_clean = str(col).strip().lower()
                if col_clean in ['الهاتف', 'رقم الهاتف', 'phone', 'mobile', 'جوال']:
                    phone_col = col
                    break

            position_col = None
            for col in df.columns:
                col_clean = str(col).strip().lower()
                if col_clean in ['الوظيفة', 'المسمى الوظيفي', 'position', 'job title', 'المنصب']:
                    position_col = col
                    break

            dept_col = None
            for col in df.columns:
                col_clean = str(col).strip().lower()
                if col_clean in ['القسم', 'department', 'dept', 'القسم/الإدارة']:
                    dept_col = col
                    break

            added = 0
            skipped = 0
            errors = []

            for idx, row in df.iterrows():
                name = str(row.get(name_col, '')).strip()
                if not name or name == 'nan':
                    skipped += 1
                    continue

                existing = Employee.query.filter_by(name=name).first()
                if existing:
                    skipped += 1
                    errors.append(f'الصف {idx + 2}: الموظف "{name}" موجود بالفعل')
                    continue

                last_employee = Employee.query.order_by(Employee.id.desc()).first()
                if last_employee and last_employee.employee_id:
                    try:
                        num = int(last_employee.employee_id.replace('EMP', ''))
                        new_num = num + 1
                    except:
                        new_num = last_employee.id + 1
                else:
                    new_num = 1
                employee_id = f'EMP{new_num:04d}'

                # قراءة البريد الإلكتروني
                email_val = None
                if email_col:
                    raw_email = str(row.get(email_col, '')).strip()
                    if raw_email and raw_email != 'nan':
                        existing_email = Employee.query.filter_by(email=raw_email).first()
                        if existing_email:
                            errors.append(f'الصف {idx + 2}: البريد "{raw_email}" مستخدم بالفعل، تم تجاهله')
                        else:
                            email_val = raw_email

                # قراءة الهاتف
                phone_val = None
                if phone_col:
                    raw_phone = str(row.get(phone_col, '')).strip()
                    if raw_phone and raw_phone != 'nan':
                        phone_val = raw_phone

                # قراءة الوظيفة
                position_val = None
                if position_col:
                    raw_position = str(row.get(position_col, '')).strip()
                    if raw_position and raw_position != 'nan':
                        position_val = raw_position

                # قراءة القسم
                dept_id = None
                if dept_col:
                    dept_name = str(row.get(dept_col, '')).strip()
                    if dept_name and dept_name != 'nan':
                        dept = Department.query.filter_by(name=dept_name).first()
                        if dept:
                            dept_id = dept.id
                        else:
                            errors.append(f'الصف {idx + 2}: القسم "{dept_name}" غير موجود، تم إضافة الموظف بدون قسم')

                employee = Employee(
                    employee_id=employee_id,
                    name=name,
                    email=email_val,
                    phone=phone_val,
                    position=position_val,
                    department_id=dept_id,
                    status='active'
                )
                db.session.add(employee)
                added += 1

            db.session.commit()

            log_action('create', 'employee', description=f'استيراد موظفين من Excel: {added} إضافة، {skipped} تخطي')

            msg = f'تم استيراد {added} موظف بنجاح'
            if skipped > 0:
                msg += f' ({skipped} تم تخطيهم)'
            flash(msg, 'success')

            if errors:
                for err in errors[:10]:  # عرض أول 10 أخطاء فقط
                    flash(err, 'warning')

            return redirect(url_for('employees.employees'))

        except Exception as e:
            flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
            return redirect(request.url)

    return render_template('import_employees.html')


@reports_bp.route('/download_import_template', endpoint='download_import_template')
@login_required()
def download_import_template():
    """تحميل قالب Excel نموذجي لاستيراد الموظفين"""
    template_data = [{
        'الاسم': 'أحمد محمد',
        'البريد': 'ahmed@example.com',
        'الهاتف': '0790123456',
        'الوظيفة': 'مهندس برمجيات',
        'القسم': 'تقنية المعلومات',
    }, {
        'الاسم': 'سارة علي',
        'البريد': 'sara@example.com',
        'الهاتف': '0790765432',
        'الوظيفة': 'محاسبة',
        'القسم': 'المالية',
    }, {
        'الاسم': 'خالد حسن',
        'البريد': '',
        'الهاتف': '',
        'الوظيفة': 'سائق',
        'القسم': '',
    }]

    df = pd.DataFrame(template_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='استيراد موظفين')

        workbook = writer.book
        worksheet = writer.sheets['استيراد موظفين']

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # تنسيق رأس الجدول
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC'),
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        worksheet.row_dimensions[1].height = 30

        # تنسيق صفوف البيانات النموذجية
        sample_font = Font(name='Arial', size=11, color='95A5A6', italic=True)
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 22
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = sample_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        # حساب عرض الأعمدة تلقائياً
        _calculate_column_widths(worksheet, df, min_width=15)

        worksheet.sheet_view.rightToLeft = True

        # ملاحظات أسفل الجدول
        note_row = len(df) + 3
        notes = [
            'ملاحظات:',
            '- عمود "الاسم" مطلوب (يمكن أيضاً استخدام: الموظف أو name)',
            '- باقي الأعمدة اختيارية',
            '- عمود "القسم" يجب أن يحتوي على اسم قسم موجود في النظام',
            '- احذف البيانات النموذجية قبل الاستيراد',
        ]
        for i, note in enumerate(notes):
            cell = worksheet.cell(row=note_row + i, column=1, value=note)
            if i == 0:
                cell.font = Font(name='Arial', bold=True, size=11, color='E74C3C')
            else:
                cell.font = Font(name='Arial', size=10, color='7F8C8D')

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employee_import_template.xlsx'
    )
